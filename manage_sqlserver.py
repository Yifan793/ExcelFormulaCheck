import math
import sqlite3

import pandas as pd
import openpyxl as pyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment
from openpyxl import load_workbook
import re
import numpy as np
from openpyxl.worksheet.cell_range import MultiCellRange

global round_valid_types_2
round_valid_types_2 = set()
global sheet_formula
global value_types
value_types = ["Text", "Integer", "Decimal", "Date", "Time", "Boolean"]
global text_invalid_types
text_invalid_types = ["ACOS", "ASIN", "ATAN", "ATAN2", "COS", "DATEDIF", "DAY", "DEGREES", "EXP", "INT", "LOG10", "MONTH", "POWER",
                      "ROUND", "SIGN", "SIN", "SQRT", "TAN", "YEAR"]
global cannot_be_null_types
cannot_be_null_types = ["SIGN"]
global datedif_params
datedif_params = ["Y", "M", "D", "MD", "YM", "YD"]
global InitPath
InitPath = './init.xlsx'
global TestCases
TestCases = "./TestCasesSqlServer.xlsx"
global AllFormulaTypeTest
AllFormulaTypeTest = "./AllFormulaTypeTest_SqlServer.xlsx"
global AfterAllFormulaTypeTest_SqlServer
AfterAllFormulaTypeTest_SqlServer = "./AfterAllFormulaTypeTest_SqlServer.xlsx"


# 查找某列所在的索引
def find_column_index(worksheet, column_value):
    idx = 1
    for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)):
        if cell == column_value:
            return idx
        idx = idx + 1


# 查找某列所在的列名
def find_column_letter(worksheet, column_value):
    return pyxl.utils.get_column_letter(find_column_index(worksheet, column_value))


def my_range(start, end):
    while start <= end:
        yield start
        start += 1


# 生成Formula、FormulaString、Type列
def generate_formula_string(sheet_value):
    init_values = {}
    for value_type in value_types:
        sheet_value_with_nan = sheet_value[value_type].dropna()
        sheet_value_with_nan.loc[len(sheet_value_with_nan)] = np.nan
        init_values[value_type] = sheet_value_with_nan

    writer = pd.ExcelWriter('./generate.xlsx', engine='xlsxwriter')

    for i in range(len(sheet_formula)):
        formula_type = sheet_formula.loc[i]["Type"]
        print("...begin generate formula string: " + formula_type)
        formula_min = sheet_formula.loc[i]["Min"]
        formula_max = sheet_formula.loc[i]["Max"]
        row_datas = []
        for num in my_range(formula_min, formula_max):
            if num == 0:
                row_data = {"Formula": formula_type,
                            "ExpectedResult": 0,
                            "FormulaString": formula_type + "()"}
                row_datas.append(row_data)
            if num == 1:
                for value_type in value_types:
                    if value_type == "Text" and text_invalid_types.__contains__(formula_type):
                        continue
                    for init_value in init_values[value_type]:
                        if isinstance(init_value, float) and math.isnan(init_value) and cannot_be_null_types.__contains__(formula_type):
                            continue
                        row_data = {"Formula": formula_type,
                                    "ExpectedResult": 0,
                                    "FormulaString": formula_type + "([" + value_type + "1])",
                                    value_type + "1": init_value}
                        row_datas.append(row_data)
            elif num == 2:
                for value_type1 in value_types:
                    if value_type1 == "Text" and text_invalid_types.__contains__(formula_type):
                        continue
                    for value_type2 in value_types:
                        if value_type2 == "Text" and text_invalid_types.__contains__(formula_type):
                            continue
                        for init_value1 in init_values[value_type1]:
                            if isinstance(init_value1, float) and math.isnan(init_value1) and cannot_be_null_types.__contains__(formula_type):
                                continue
                            for init_value2 in init_values[value_type2]:
                                if isinstance(init_value2, float) and math.isnan(init_value2) and cannot_be_null_types.__contains__(formula_type):
                                    continue
                                if formula_type == "ROUND":
                                    if not isinstance(init_value2, float):
                                        continue
                                    # 可以是nan，或者0-15之间的数字
                                    if not math.isnan(init_value2) and (init_value2 < 0 or init_value2 > 15):
                                        continue
                                    round_valid_types_2.add(value_type2)
                                    row_data_round = {"Formula": formula_type,
                                                      "ExpectedResult": 0,
                                                      "FormulaString": formula_type + "([" + value_type1 + "1], [" + value_type2 + "2])",
                                                      value_type1 + "1": init_value1,
                                                      value_type2 + "2": init_value2 if math.isnan(init_value2) else int(init_value2)}
                                    row_datas.append(row_data_round)
                                elif formula_type == "POWER":
                                    # POWER(Negative, Decimal)报错 NAN，相当于把这个负数开根号
                                    if isinstance(init_value1, float) and isinstance(init_value2, float) and init_value1 < 0 and init_value2 % 1 != 0:
                                        continue
                                    row_data_power = {"Formula": formula_type,
                                                      "ExpectedResult": 0,
                                                      "FormulaString": formula_type + "([" + value_type1 + "1], [" + value_type2 + "2])",
                                                      value_type1 + "1": init_value1,
                                                      value_type2 + "2": init_value2}
                                    row_datas.append(row_data_power)
                                else:
                                    row_data = {"Formula": formula_type,
                                                "ExpectedResult": 0,
                                                "FormulaString": formula_type + "([" + value_type1 + "1], [" + value_type2 + "2])",
                                                value_type1 + "1": init_value1,
                                                value_type2 + "2": init_value2}
                                    row_datas.append(row_data)
            elif num == 3:
                for value_type1 in value_types:
                    if value_type1 == "Text" and text_invalid_types.__contains__(formula_type):
                        continue
                    for value_type2 in value_types:
                        if value_type2 == "Text" and text_invalid_types.__contains__(formula_type):
                            continue
                        if formula_type == "DATEDIF":
                            for init_value1 in init_values[value_type1]:
                                if isinstance(init_value1, float) and math.isnan(
                                        init_value1) and cannot_be_null_types.__contains__(formula_type):
                                    continue
                                for init_value2 in init_values[value_type2]:
                                    if isinstance(init_value2, float) and math.isnan(
                                            init_value2) and cannot_be_null_types.__contains__(formula_type):
                                        continue
                                    for datedif_param in datedif_params:
                                        row_data_datedif = {"Formula": formula_type,
                                                            "ExpectedResult": 0,
                                                            "FormulaString": formula_type + "([" + value_type1 + "1], [" +
                                                                             value_type2 + "2], \"" + datedif_param + "\")",
                                                            value_type1 + "1": init_value1,
                                                            value_type2 + "2": init_value2}
                                        row_datas.append(row_data_datedif)
                        else:
                            for value_type3 in value_types:
                                if value_type3 == "Text" and text_invalid_types.__contains__(formula_type):
                                    continue
                                for init_value1 in init_values[value_type1]:
                                    if isinstance(init_value1, float) and math.isnan(init_value1) and cannot_be_null_types.__contains__(formula_type):
                                        continue
                                    for init_value2 in init_values[value_type2]:
                                        if isinstance(init_value2, float) and math.isnan(init_value2) and cannot_be_null_types.__contains__(formula_type):
                                            continue
                                        for init_value3 in init_values[value_type3]:
                                            if isinstance(init_value3, float) and math.isnan(init_value3) and cannot_be_null_types.__contains__(formula_type):
                                                continue
                                            row_data = {"Formula": formula_type,
                                                        "ExpectedResult": 0,
                                                        "FormulaString": formula_type + "([" + value_type1 + "1], ["
                                                                         + value_type2 + "2], [" + value_type3 + "3])",
                                                        value_type1 + "1": init_value1,
                                                        value_type2 + "2": init_value2,
                                                        value_type3 + "3": init_value3}
                                            row_datas.append(row_data)
        df = pd.DataFrame(row_datas)
        df.to_excel(writer, sheet_name=formula_type, index=False)
    writer._save()


# 填充ExpectedResult数据
def generate_expected_result():
    types_to_format = {"Text": '@',
                       "Integer": '0',
                       "Decimal": '0.00000',
                       "Date": 'yyyy/m/d',
                       "Time": 'h:mm:ss',
                       "Boolean": 'General'}
    workbook = load_workbook('./generate.xlsx')
    for i in range(len(sheet_formula)):
        formula_type = sheet_formula.loc[i]["Type"]
        print("...begin generate expected result: " + formula_type)
        formula_max = sheet_formula.loc[i]["Max"]
        worksheet = workbook[formula_type]

        if formula_type == "SIGN":
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        cell.value = 0

        # 修改单元格格式
        for idx in range(int(formula_max)):
            if formula_type == "DATEDIF" and idx == 2:
                continue
            for value_type in value_types:
                if value_type == "Text" and text_invalid_types.__contains__(formula_type):
                    continue
                if formula_type == "ROUND" and idx == 1 and not round_valid_types_2.__contains__(value_type):
                    continue
                column_letter = find_column_letter(worksheet, value_type + str(idx + 1))
                for cell in worksheet[column_letter]:
                    cell.number_format = types_to_format[value_type]
        # 填充ExpectedResult
        column_name_to_letter = {}
        column_expected_result_letter = find_column_letter(worksheet, "ExpectedResult")
        column_formula_string_letter = find_column_letter(worksheet, "FormulaString")
        for col_idx in range(len(worksheet['A'])):
            if col_idx == 0:
                continue
            col_idx_string = str(col_idx + 1)
            formula_string = worksheet[column_formula_string_letter + col_idx_string].value
            results = re.findall(r"\[.*?\]", formula_string)
            expected_result = formula_string
            for result in results:
                if result[1:-1] in column_name_to_letter:
                    formula_letter = column_name_to_letter[result[1:-1]]
                else:
                    formula_letter = find_column_letter(worksheet, result[1:-1])
                    column_name_to_letter[result[1:-1]] = formula_letter
                expected_result = expected_result.replace(result, formula_letter + col_idx_string)
            worksheet[column_expected_result_letter + col_idx_string] = "=" + expected_result
    workbook.save(TestCases)


def generate_test_file():
    global sheet_formula
    sheet_formula = pd.read_excel(InitPath, sheet_name="Formula")
    sheet_value = pd.read_excel(InitPath, sheet_name="Value")

    generate_formula_string(sheet_value)
    generate_expected_result()


def after_sql_to_excel():
    global sheet_formula
    sheet_formula = pd.read_excel(InitPath, sheet_name="Formula")
    workbook = load_workbook(AllFormulaTypeTest, data_only=True)
    excel_file = pd.ExcelFile(AllFormulaTypeTest)
    sheet_names = excel_file.sheet_names
    for i in range(len(sheet_formula)):
        formula_type = sheet_formula.loc[i]["Type"]
        print("after_sql_to_excel: " + formula_type)
        if formula_type not in sheet_names:
            continue
        df = pd.read_excel(AllFormulaTypeTest, sheet_name=formula_type)
        worksheet = workbook[formula_type]

        fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        fill_blue = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
        fill_green = PatternFill(start_color="92D050", end_color="92D050", fill_type='solid')

        # 填充ExpectedResult和颜色
        column_name_to_letter = {}
        column_same_letter = find_column_letter(worksheet, "IsExcelSameAsDataBase")
        column_expected_result_letter = find_column_letter(worksheet, "ExpectedResult")
        column_formula_string_letter = find_column_letter(worksheet, "FormulaString")
        for row_idx, row in df.iterrows():
            row_idx_string = str(row_idx + 2)
            if row['IsExcelSameAsDataBase'] == 'false' or not row['IsExcelSameAsDataBase']:
                worksheet[column_same_letter + row_idx_string].fill = fill_yellow
            elif row['IsExcelSameAsDataBase'] == "invalid":
                worksheet[column_same_letter + row_idx_string].fill = fill_green
            formula_string = worksheet[column_formula_string_letter + row_idx_string].value
            replaced_formula_string = formula_string.replace('(', '_').replace(')', '_').replace('[', '_').replace(']', '_').replace(',', '_').replace('\"', '_').replace(" ", "")
            if replaced_formula_string in column_name_to_letter:
                replaced_formula_letter = column_name_to_letter[replaced_formula_string]
            else:
                replaced_formula_letter = find_column_letter(worksheet, replaced_formula_string)
                column_name_to_letter[replaced_formula_string] = replaced_formula_letter
            worksheet[replaced_formula_letter + row_idx_string].fill = fill_blue
            results = re.findall(r"\[.*?\]", formula_string)
            expected_result = formula_string
            for result in results:
                if result[1:-1] in column_name_to_letter:
                    formula_letter = column_name_to_letter[result[1:-1]]
                else:
                    formula_letter = find_column_letter(worksheet, result[1:-1])
                    column_name_to_letter[result[1:-1]] = formula_letter
                # worksheet[formula_letter + row_idx_string].fill = fill_blue
                # 由于最后在首行添加了空白行，所以expected_result的row_idx需要加1
                expected_result = expected_result.replace(result, formula_letter + str(int(row_idx_string) + 1))
            worksheet[column_expected_result_letter + row_idx_string] = "=" + expected_result
            worksheet[column_expected_result_letter + row_idx_string].fill = fill_blue

        # 插入空白行填写结论
        worksheet.insert_rows(1)

    # 保存修改后的Excel文件
    workbook.save(AfterAllFormulaTypeTest_SqlServer)


def add_conclusion():
    global sheet_formula
    sheet_formula = pd.read_excel(InitPath, sheet_name="Formula")
    workbook1 = load_workbook('./AllFormulaTypeTestResultWithConclusion_SqlServer.xlsx')
    workbook2 = load_workbook(AfterAllFormulaTypeTest_SqlServer)
    excel_file = pd.ExcelFile(AfterAllFormulaTypeTest_SqlServer)
    sheet_names = excel_file.sheet_names
    alignment = Alignment(vertical='top', wrapText=True)
    for i in range(len(sheet_formula)):
        formula_type = sheet_formula.loc[i]["Type"]
        print("add_conclusion: " + formula_type)
        if formula_type not in sheet_names:
            continue
        sheet1 = workbook1[formula_type]
        sheet2 = workbook2[formula_type]
        sheet2.row_dimensions[1].height = sheet1.row_dimensions[1].height
        dealt_merged_ranges = MultiCellRange()
        merged_ranges = sheet1.merged_cells.ranges
        for cell in sheet1[1]:
            if cell.coordinate in sheet1.merged_cells:
                should_skip = False
                for dealt_range in dealt_merged_ranges:
                    if cell.coordinate in dealt_range:
                        should_skip = True
                        break
                if should_skip:
                    continue
                for merged_range in merged_ranges:
                    if cell.coordinate in merged_range:
                        dealt_merged_ranges.add(merged_range)
                        start_cell = merged_range.start_cell
                        merged_value = start_cell.value
                        sheet2.merge_cells(str(merged_range))
                        target_cell = sheet2[start_cell.coordinate]
                        target_cell.value = merged_value
                        target_cell.alignment = alignment
            else:
                target_cell = sheet2.cell(row=1, column=cell.column)
                target_cell.value = cell.value
                target_cell.alignment = alignment
    workbook2.save('./After_AllFormulaTypeTestResultWithConclusion_SqlServer.xlsx')


def generate_expected_result_valid():
    global sheet_formula
    sheet_formula = pd.read_excel(InitPath, sheet_name="Formula")
    workbook = load_workbook('./TestCasesWithValidDatas.xlsx')
    for i in range(len(sheet_formula)):
        formula_type = sheet_formula.loc[i]["Type"]
        print("...begin generate expected result: " + formula_type)
        worksheet = workbook[formula_type]

        # 填充ExpectedResult
        column_name_to_letter = {}
        column_expected_result_letter = find_column_letter(worksheet, "ExpectedResult")
        column_formula_string_letter = find_column_letter(worksheet, "FormulaString")
        for col_idx in range(len(worksheet['A'])):
            if col_idx == 0:
                continue
            col_idx_string = str(col_idx + 1)
            formula_string = worksheet[column_formula_string_letter + col_idx_string].value
            results = re.findall(r"\[.*?\]", formula_string)
            expected_result = formula_string
            for result in results:
                if result[1:-1] in column_name_to_letter:
                    formula_letter = column_name_to_letter[result[1:-1]]
                else:
                    formula_letter = find_column_letter(worksheet, result[1:-1])
                    column_name_to_letter[result[1:-1]] = formula_letter
                expected_result = expected_result.replace(result, formula_letter + col_idx_string)
            worksheet[column_expected_result_letter + col_idx_string] = "=" + expected_result
    workbook.save('./After_TestCasesWithValidDatas.xlsx')


if __name__ == '__main__':
    round_valid_types_2 = set()
    generate_test_file()
    # after_sql_to_excel()
    # add_conclusion()
    # generate_expected_result_valid()
